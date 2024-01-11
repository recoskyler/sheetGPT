#!/usr/bin/python

from openpyxl import load_workbook
from os.path import isfile, exists
from os import remove
from os.path import splitext
from shutil import copyfile

__all__ = ["load_input_workbook", "create_output_book", "generate_result_path"]

def load_input_workbook(file_path: str):
    wb = None

    if not exists(file_path) or not isfile(file_path):
        raise Exception("File does not exist or invalid\n")

    try:
        wb = load_workbook(file_path, read_only=True)
    except Exception as e:
        print(e)
        raise Exception("Failed to load workbook. Is the file format correct (xlsx/xlsm/xltx/xltm)?\n")

    return wb

def create_output_book(file_path: str, result_path: str, overwrite=False):
    result_book = None

    try:
        copyfile(file_path, result_path)
        result_book = load_workbook(result_path)
    except Exception as e:
        print(e)
        raise Exception("\nFailed to create output worksheet.\n")

    return result_book

def generate_result_path(file_path: str) -> str:
    pre, ext = splitext(file_path)
    return pre + "_output" + ext
