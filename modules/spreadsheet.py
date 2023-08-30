#!/usr/bin/python

from openpyxl import load_workbook
from os.path import isfile, exists
from os import remove
from shutil import copyfile

__all__ = ["load_input_workbook", "create_result_book"]

def load_input_workbook(file_path: str):
    wb = None

    if not exists(file_path) or not isfile(file_path):
        print("File does not exist or invalid\n")
        return False

    try:
        print("\nLoading workbook...")

        wb = load_workbook(file_path, read_only=True)

        print("Loaded workbook\n")
    except Exception as e:
        print(e)
        print("Failed to load workbook. Is the file format correct (.xlsx or .xlsm)?\n")
        return False

    return wb

def create_result_book(file_path: str, result_path: str):
    overwrite = False
    result_book = None

    if exists(result_path) and isfile(result_path):
        while True:
            res = input("Output worksheet already exists. Would you like to delete the old worksheet and create it again? (y/n)  ").strip().lower()

            if res == "y":
                overwrite = True

                print("Deleting old output file...\n")

                try:
                    remove(result_path)
                except:
                    print("Failed to delete old output worksheet.\n")
                    return False

                break
            elif res == "n":
                print("Using the old file...")
                break

            print("\nPlease enter a valid choice.\n")

    try:
        if not exists(result_path) or not isfile(result_path) or overwrite:
            print("Creating output worksheet...")

            copyfile(file_path, result_path)

        result_book = load_workbook(result_path)

        print("Loaded output worksheet: " + result_path)
    except Exception as e:
        print("\nFailed to create output worksheet.\n")
        print(e)

        if exists(result_path) and isfile(result_path):
            print("Removing invalid output file...")

            try:
                remove(result_path)
            except:
                print("Failed to remove invalid output file")

        return False

    return result_book