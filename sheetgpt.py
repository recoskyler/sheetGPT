#!/usr/bin/python

import re
import os
import openai
import openpyxl
import signal
import shutil


if __name__ != '__main__':
    exit()

# Constants

DEFAULT_SYSTEM_PROMPT = "You are a helpful assistant. You give only the answer, without forming a sentence. If you are not sure, try guessing. If you are still unsure about the answer, output '?'. If you don't know the answer, or if you cannot give a correct answer output '?'."

MODELS = ["gpt-3.5-turbo", "gpt-3.5-turbo-16k", "gpt-4", "gpt-4-32k"]

INPUTS_REGEX = "(^([0-9]+)([, ]([0-9]+))*$)|(^([a-zA-Z]+)([, ]([a-zA-Z]+))*$)"

RESULT_REGEX = "^([a-zA-Z]+)([0-9]+)$"

SAVE_INTERVAL = 25 # Save after every 25 processed items

# Variables

file_path = ""
result_path = ""
api_key = ""
model = ""
sheet_name = "Sheet 1"
system_prompt = ""
prompt = ""
result_row = 1
result_col = ""
inputs = ""
workbook = None
worksheet = None
result_book = None
result_sheet = None
by_row = False # Output the result to the next row?
by_col = False # Output the result to the next column?
input_pos = ""
cache = dict()
limit = 0
processed = 0

# Functions

def handler(signum, frame):
    if workbook != None:
        workbook.close()

    if result_book != None:
        print("\nSaving file...")

        result_book.save(result_path)
        result_book.close()
    else:
        print("\n\nExiting without saving...")

    exit(0)

def load_workbook():
    global workbook
    global file_path

    if not os.path.exists(file_path) or not os.path.isfile(file_path):
        print("File does not exist or invalid\n")
        return False

    try:
        print("\nLoading worksheet...")

        workbook = openpyxl.load_workbook(file_path, read_only=True)

        print("Loaded worksheet\n")
    except:
        print("Failed to load worksheet. Is the file format correct (.xlsx or .xlsm)?\n")
        return False

    return True

def create_result_sheet():
    global result_book
    global result_path
    global file_path

    pre, ext = os.path.splitext(file_path)

    result_path = pre + "_output" + ext
    overwrite = False

    if os.path.exists(result_path) and os.path.isfile(result_path):
        while True:
            res = input("Output worksheet already exists. Would you like to delete the old worksheet and create it again? (y/n)  ").strip().lower()

            if res == "y":
                overwrite = True

                print("Deleting old output file...\n")

                try:
                    os.remove(result_path)
                except:
                    print("Failed to delete old output worksheet.\n")
                    return False

                break
            elif res == "n":
                print("Using the old file...")
                break

            print("\nPlease enter a valid choice.\n")

    try:
        if not os.path.exists(result_path) or not os.path.isfile(result_path) or overwrite:
            print("Creating output worksheet...")

            shutil.copyfile(file_path, result_path)

        result_book = openpyxl.load_workbook(result_path)

        print("Created output worksheet: " + result_path)
    except Exception as e:
        print("\nFailed to create output worksheet.\n")
        print(e)
        return False

    return True

def get_inputs():
    global file_path
    global api_key
    global model
    global sheet_name
    global system_prompt
    global prompt
    global result_book
    global result_sheet
    global result_row
    global result_col
    global inputs
    global worksheet
    global workbook
    global by_row
    global by_col
    global input_pos
    global limit

    while True:
        file_path = input("\nEnter the full or relative file path (.xlsx/.xlsm): ").strip()

        if load_workbook() and create_result_sheet():
            break

    if len(workbook.sheetnames) == 0:
        print("No sheets found")
        exit(0)

    while True:
        print("\nSheets in the worksheet:\n")

        index = 1

        for sheet in workbook.sheetnames:
            print(str(index) + ") " + sheet)
            index = index + 1

        choice = input("\nChoose sheet (press ENTER to select the first sheet): ").strip()

        if choice == "":
            sheet_name = workbook.sheetnames[0]

            break
        elif choice.isnumeric() and int(choice) >= 1 and int(choice) <= len(workbook.sheetnames):
            sheet_name = workbook.sheetnames[int(choice) - 1]

            print("\nUsing the sheet with name: ")
            print(sheet_name)
            print("\n")

            break

        print("\nPlease choose a valid sheet")

    worksheet = workbook[sheet_name]
    result_sheet = result_book[sheet_name]

    while True:
        api_key = input("\nEnter your OpenAI API key: ").strip()

        if api_key != "":
            openai.api_key = api_key
            break

        print("\nPlease enter a valid API key\n")

    while True:
        choice = input("\nChoose a ChatGPT model:\n\n1) gpt-3.5-turbo\n2) gpt-3.5-turbo-16k (default)\n3) gpt-4\n4) gpt-4-32k\n\nChoice (1-4, ENTER for default): ").strip()

        if choice == "":
            model = MODELS[1]
            print("\nUsing the default model\n")
            break

        if choice.isnumeric() and int(choice) > 0 and int(choice) < 5:
            model = MODELS[int(choice) - 1]
            print("\nUsing the model: ")
            print(model)
            break

        print("\nPlease choose a valid model\n")

    system_prompt = input("\nEnter a system prompt to fine-tune the response (press ENTER for default):\n\n").strip()

    if system_prompt == "":
        system_prompt = DEFAULT_SYSTEM_PROMPT

        print("\nUsing the default prompt:\n")
        print(system_prompt)
        print("\n")

    while True:
        prompt = input("\nEnter a prompt to be used with each request. Use zero-based '$0' placeholder(s) to insert inputs (i.e. Assuming you have entered 'a,b' as the inputs, and '2' as the starting row: 'Something $0 and $0, $1' will be converted to 'Something <VALUE OF A2> and <VALUE OF A2>, <VALUE OF B2>')\n\n").strip()

        if prompt != "":
            break

        print("\nPlease enter a valid prompt\n")

    while True:
        inputs = input("\nEnter one or more input columns (A-Z)/rows (1-9) separated by a space ('1 3') or a comma ('1,3')  character:\n\n").strip().strip(",")

        if inputs != "" and re.search(INPUTS_REGEX, inputs) != None:
            inputs = re.split(",| ", inputs.upper())
            break

        print("\nPlease enter a valid input. You must choose either row(s) (i.e. '2' or '4,2,0,6,9'), or column(s) (i.e. 'd' or 'p,r,n,d'). Not both (i.e. '1,a,c,4').\n")

    while True:
        question = "\nEnter the starting row for the inputs (i.e. '2'): "

        if inputs[0].isnumeric():
            question = "\nEnter the starting column for the inputs (i.e. 'C'): "

        choice = input(question).strip()

        if inputs[0].isnumeric() and choice.isalpha():
            input_pos = choice.upper()
            break

        if inputs[0].isalpha() and choice.isnumeric():
            input_pos = choice
            break

        if inputs[0].isnumeric():
            print("\nPlease enter a valid column (must be a character)\n")
        else:
            print("\nPlease enter a valid row (must be a number)\n")

    while True:
        result_col = input("\nEnter the starting column for the results (i.e. 'B'): ").strip().upper()

        if result_col != "" and result_col.isalpha():
            break

        print("\nPlease enter a valid column.\n")

    while True:
        result_row = input("\nEnter the starting row for the results (i.e. '2'): ").strip()

        if result_row != "" and result_row.isnumeric():
            result_row = result_row
            break

        print("\nPlease enter a valid row.\n")

    result_placement = result_col + str(result_row)

    print("\n\nStarting point for results: ")
    print(result_placement)
    print("\n")

    while True:
        choice = input("\nShould the result be output to the next row, or column? (type 'r' for next row, 'c' for next column) ").strip()

        if choice == "r" or choice == "c":
            by_col = choice == "c"
            by_row = choice == "r"
            break

        print("\nPlease enter a valid choice ('r' or 'c')\n")

    while True:
        limit = input("\nEnter a limit for number of items to be processed (type 0 to remove limit): ").strip()

        if limit.isnumeric() and int(limit) >= 0:
            limit = int(limit)
            break

        print("\nPlease enter a valid limit.\n")

def get_next_column(col):
    alphabet = "abcdefghijklmnopqrstuvwxyz".upper().split()

    if alphabet.index(col.upper()) + 1 < len(alphabet):
        return col.upper()[:-1] + alphabet[alphabet.index(col.upper()[-1]) + 1]

    return col.upper()[:-1] + alphabet[0]

def get_answer(prompt):
    global model
    global system_prompt
    global cache
    global processed

    prompt_hash = str(hash(prompt))

    if prompt_hash in cache.keys():
        print("Using cached answer...")
        return cache[prompt_hash]

    try:
        print("Getting answer from ChatGPT...")

        res = openai.ChatCompletion.create(model=model, messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": prompt}]).choices[0].message.content

        processed = processed + 1
    except Exception as e:
        print("Failed to get answer: ")
        print(e)
        return ""

    cache[prompt_hash] = res

    return res

def save_progress():
    try:
        print("\nSaving file...")

        result_book.save(result_path)

        print("\nResults saved to: " + result_path)
    except Exception as e:
        print("\nAn error occurred while saving\n\n")
        print(e)

def process_worksheet():
    global worksheet
    global result_sheet
    global by_row
    global by_col
    global input_pos
    global prompt
    global result_book
    global result_sheet
    global result_row
    global result_col
    global inputs
    global limit
    global processed

    ended = False
    item_no = 1
    processed = 0

    print("Processing... Press CTRL + C to cancel")

    if limit == 0:
        print("\nWARNING: No limit set. Results won't be saved until the end\n")
    else:
        print("\nLimiting to " + str(limit) + " items...\n")

    while limit == 0 or processed <= limit:
        if processed != 0 and processed % SAVE_INTERVAL == 0: save_progress()

        print("Processing item " + str(processed) + "/" + str(item_no) + "...")

        input_values = []

        for item in inputs:
            coordinates = ""

            if item.isalpha():
                coordinates = item.upper() + str(input_pos)
            else:
                coordinates = input_pos.upper() + item

            # print("Getting input at coordinate " + coordinates)

            value = worksheet[coordinates].value.strip()

            if value == "":
                ended = True
                break

            input_values.append(value)
            # print("Input " + str(item_no) + "." + str(len(input_values)) + ": " + value)

        if ended:
            break

        current_prompt = prompt
        index = int(re.search("(\$)([0-9]+)", current_prompt).group()[1:])

        while re.search("(\$)([0-9]+)", current_prompt) != None:
            match = re.search("(\$)([0-9]+)", current_prompt).group()
            index = int(match[1:])

            if index < len(input_values):
                current_prompt = re.sub("(\$)([0-9]+)", input_values[index], current_prompt, 1)
            else:
                print("Found invalid placeholder, replacing with empty string: " + match)

                current_prompt = re.sub("(\$)([0-9]+)", "", current_prompt, 1)

                break

        if result_sheet[result_col + result_row].value == None or str(result_sheet[result_col + result_row].value).strip() == "":
            print("Prompt: " + current_prompt)

            res = get_answer(current_prompt)
            result_sheet[result_col + result_row] = res

            print("Answer: " + res)
        else:
            print("Already has result. Skipping...")

            prompt_hash = str(hash(current_prompt))

            if prompt_hash not in cache.keys():
                print("Caching answer...")
                cache[prompt_hash] = result_sheet[result_col + result_row].value.strip()

        if by_row:
            result_row = str(int(result_row) + 1)
        else:
            result_col = get_next_column(result_col)

        if input_pos.isnumeric():
            input_pos = str(int(input_pos) + 1)
        else:
            input_pos = get_next_column(input_pos)

        item_no = item_no + 1

    print("Done!")

def clear():
    if os.name == 'nt':
        _ = os.system('cls')
    else:
        _ = os.system('clear')

# Ctrl+C handling

signal.signal(signal.SIGINT, handler)

# Program

clear()

print("\n\nSheetGPT v1.0.0 by recoskyler\n\n")

get_inputs()

try:
    process_worksheet()
except KeyboardInterrupt:
    print("Exiting...")
except Exception as e:
    print("\nAn error occurred\n\n")
    print(e)
finally:
    if workbook != None:
        workbook.close()

    if result_book != None:
        print("\nSaving file...")

        result_book.save(result_path)
        result_book.close()

        print("\nResults saved to: " + result_path)
    else:
        print("\n\nExiting without saving...")

    print("\n======= Adios, Cowboy =======")
